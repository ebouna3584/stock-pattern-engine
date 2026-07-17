"""
Quick Analyze endpoints — fetch OHLCV from Yahoo Finance, run the full pipeline.

POST /quick_analyze   → JSON analysis results
GET  /prefilled_template → pre-filled Excel download
"""

import io
import logging
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scheduler.price_fetcher import fetch_history
from core.engine import run_analysis
from api.config import settings
from models.schemas import AnalysisResponse
from auth.dependencies import get_current_user
from db.models import User

logger = logging.getLogger(__name__)
router = APIRouter()

_COLUMNS = [
    "template_version", "ticker", "timeframe", "date",
    "open", "high", "low", "close", "volume",
    "RSI", "MACD", "MACD_signal", "MACD_histogram",
    "EMA_20", "EMA_50", "SMA_200", "Bollinger_upper", "Bollinger_lower", "ATR",
]


class QuickAnalyzeRequest(BaseModel):
    ticker: str
    buy_date: str          # YYYY-MM-DD — start of the analysis window
    end_date: Optional[str] = None  # YYYY-MM-DD — defaults to today


@router.post("/quick_analyze", response_model=AnalysisResponse)
async def quick_analyze(request: QuickAnalyzeRequest, user: User = Depends(get_current_user)) -> AnalysisResponse:
    """
    Fetch OHLCV data from Yahoo Finance and run the full pattern engine.
    No CSV upload required — just supply a ticker and buy date.

    **Disclaimer**: Results are probabilistic technical analysis only.
    Past patterns do not guarantee future results.
    """
    ticker   = request.ticker.upper().strip()
    end_date = request.end_date or str(date.today())

    df = fetch_history(ticker, start_date=request.buy_date, end_date=end_date)

    if df is None or df.empty:
        raise HTTPException(status_code=422, detail=f"No data returned for '{ticker}'. Check the ticker symbol.")

    if len(df) < settings.MIN_ROWS:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Only {len(df)} trading days found for '{ticker}' from {request.buy_date} to {end_date}. "
                f"At least {settings.MIN_ROWS} rows are required for pattern detection. "
                "Try an earlier buy date."
            ),
        )

    response = run_analysis(df=df)

    # Cache result so report download buttons work
    try:
        from api.endpoints.report import store_result
        store_result(response.session_id, response)
    except Exception:
        pass

    return response


@router.get("/prefilled_template")
async def prefilled_template(
    ticker: str,
    buy_date: str,
    end_date: Optional[str] = None,
    user: User = Depends(get_current_user),
):
    """
    Download a pre-filled Excel file with real OHLCV data from Yahoo Finance.
    User only needs to enter ticker + buy date — all other data is auto-filled.
    The downloaded file can be reviewed, edited, and re-uploaded via the CSV upload flow.
    """
    ticker   = ticker.upper().strip()
    end_date = end_date or str(date.today())

    df = fetch_history(ticker, start_date=buy_date, end_date=end_date)

    if df is None or df.empty:
        raise HTTPException(status_code=422, detail=f"No data returned for '{ticker}'. Check the ticker symbol.")

    buf = io.BytesIO()
    _write_prefilled_excel(df, ticker, buf)
    buf.seek(0)

    filename = f"{ticker}_prefilled_{buy_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_prefilled_excel(df, ticker: str, buf: io.BytesIO) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ticker

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    center   = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, col_name in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 12)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    row_fill_a = PatternFill("solid", fgColor="F8FAFC")
    row_fill_b = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = row_fill_a if row_idx % 2 == 0 else row_fill_b
        for col_idx, col_name in enumerate(_COLUMNS, 1):
            val = row.get(col_name)
            if hasattr(val, 'item'):
                val = val.item()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=9)
            cell.fill      = fill
            cell.alignment = center

    wb.save(buf)
