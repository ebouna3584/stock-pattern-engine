"""
Excel Template Generator
========================
Generates a standardised .xlsx download that users fill in and re-upload as CSV.

Usage:
    from templates.excel_generator import build_template_workbook
    wb = build_template_workbook()
    wb.save("StockPatternTemplate_v1.0.0.xlsx")
"""

import io
import numpy as np
import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ─── Constants ────────────────────────────────────────────────────────────────

TEMPLATE_VERSION = "1.0.0"
FILENAME = f"StockPatternTemplate_v{TEMPLATE_VERSION}.xlsx"

COLUMNS = [
    "template_version",
    "ticker",
    "timeframe",
    "date",
    "open",
    "high",
    "low",
    "close",
    "volume",
    "RSI",
    "MACD",
    "MACD_signal",
    "MACD_histogram",
    "EMA_20",
    "EMA_50",
    "SMA_200",
    "Bollinger_upper",
    "Bollinger_lower",
    "ATR",
]

# Fixed column widths (characters)
COLUMN_WIDTHS = {
    "template_version": 18,
    "ticker":           10,
    "timeframe":        12,
    "date":             14,
    "open":             12,
    "high":             12,
    "low":              12,
    "close":            12,
    "volume":           16,
    "RSI":              10,
    "MACD":             12,
    "MACD_signal":      14,
    "MACD_histogram":   16,
    "EMA_20":           12,
    "EMA_50":           12,
    "SMA_200":          12,
    "Bollinger_upper":  16,
    "Bollinger_lower":  16,
    "ATR":              10,
}

TIMEFRAME_OPTIONS = ["1D", "4H", "1H", "15M"]

# ─── Styles ───────────────────────────────────────────────────────────────────

HEADER_FONT       = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL       = PatternFill("solid", fgColor="1F3864")   # dark navy
EXAMPLE_FILL      = PatternFill("solid", fgColor="F2F2F2")   # light gray
REQUIRED_FILL     = PatternFill("solid", fgColor="D6E4F7")   # light blue — required cols
INSTRUCTION_TITLE = Font(name="Calibri", bold=True, size=14, color="1F3864")
INSTRUCTION_HEAD  = Font(name="Calibri", bold=True, size=11)
INSTRUCTION_BODY  = Font(name="Calibri", size=10)
WARN_FONT         = Font(name="Calibri", bold=True, size=10, color="C00000")
CENTER            = Alignment(horizontal="center", vertical="center", wrap_text=False)
WRAP              = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Number formats
FMT_PRICE   = '0.00'            # 2 dp, no commas → exports cleanly to CSV
FMT_VOLUME  = '0'               # plain integer, no commas → exports cleanly to CSV
FMT_DATE    = 'YYYY-MM-DD'
FMT_PERCENT = '0.00'

# Columns that use the price format
PRICE_COLS = {
    "open", "high", "low", "close",
    "RSI", "MACD", "MACD_signal", "MACD_histogram",
    "EMA_20", "EMA_50", "SMA_200",
    "Bollinger_upper", "Bollinger_lower", "ATR",
}

# ─── Example Data ─────────────────────────────────────────────────────────────

def _generate_ticker_rows(ticker: str, start_close: float, atr_base: float,
                           volume_base: int, seed: int, n: int = 80) -> list:
    rng = np.random.default_rng(seed)
    rows = []
    d = date(2024, 1, 2)
    close = start_close
    closes_hist = []

    for i in range(n):
        # advance to next business day
        while d.weekday() >= 5:
            d += timedelta(days=1)

        # random walk close
        ret = rng.normal(0.0003, 0.015)
        close = round(close * (1 + ret), 2)

        atr = round(atr_base * rng.uniform(0.85, 1.25), 2)
        high = round(close + rng.uniform(0.2, 1.0) * atr, 2)
        low  = round(close - rng.uniform(0.2, 1.0) * atr, 2)
        open_ = round(low + rng.uniform(0, 1) * (high - low), 2)
        volume = int(volume_base * rng.uniform(0.7, 1.5))

        closes_hist.append(close)
        ch = np.array(closes_hist)

        # RSI (simple)
        if len(ch) >= 14:
            deltas = np.diff(ch[-15:])
            gains = np.where(deltas > 0, deltas, 0)
            losses = np.where(deltas < 0, -deltas, 0)
            avg_g = np.mean(gains) if np.mean(gains) > 0 else 1e-9
            avg_l = np.mean(losses) if np.mean(losses) > 0 else 1e-9
            rsi = round(100 - 100 / (1 + avg_g / avg_l), 1)
        else:
            rsi = round(50 + rng.normal(0, 5), 1)

        # EMA helper
        def ema(arr, span):
            s = pd.Series(arr)
            return float(s.ewm(span=span, adjust=False).mean().iloc[-1])

        ema20  = round(ema(ch, 20), 2)
        ema50  = round(ema(ch, 50), 2)
        sma200 = round(float(np.mean(ch[-200:])), 2)

        std20 = float(np.std(ch[-20:])) if len(ch) >= 20 else atr * 0.5
        bb_upper = round(ema20 + 2 * std20, 2)
        bb_lower = round(ema20 - 2 * std20, 2)

        ema12 = round(ema(ch, 12), 2)
        ema26 = round(ema(ch, 26), 2)
        macd_val = round(ema12 - ema26, 4)

        # simple MACD signal/histogram approximation
        if len(ch) >= 9:
            macd_hist_arr = [ema(np.array(closes_hist[:j+1]), 12) - ema(np.array(closes_hist[:j+1]), 26) for j in range(max(0, len(closes_hist)-9), len(closes_hist))]
            macd_signal = round(float(pd.Series(macd_hist_arr).ewm(span=9, adjust=False).mean().iloc[-1]), 4)
        else:
            macd_signal = round(macd_val * 0.8, 4)
        macd_histogram = round(macd_val - macd_signal, 4)

        rows.append([
            TEMPLATE_VERSION, ticker, "1D", d.strftime("%Y-%m-%d"),
            open_, high, low, close, volume,
            rsi, macd_val, macd_signal, macd_histogram,
            ema20, ema50, sma200, bb_upper, bb_lower, atr
        ])
        d += timedelta(days=1)

    return rows


EXAMPLE_ROWS = (
    _generate_ticker_rows("AAPL", start_close=185.0, atr_base=2.8, volume_base=55_000_000, seed=1) +
    _generate_ticker_rows("TSLA", start_close=250.0, atr_base=7.5, volume_base=100_000_000, seed=2) +
    _generate_ticker_rows("NVDA", start_close=495.0, atr_base=13.5, volume_base=45_000_000, seed=3)
)


# ─── Builder ──────────────────────────────────────────────────────────────────

def build_template_workbook() -> Workbook:
    wb = Workbook()
    _build_template_sheet(wb)
    _build_instructions_sheet(wb)
    return wb


def build_template_bytes() -> bytes:
    """Return the workbook serialised to bytes (for streaming HTTP response)."""
    wb = build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Sheet 1: TEMPLATE ────────────────────────────────────────────────────────

def _build_template_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "TEMPLATE"

    # ── Headers ───────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    # ── Freeze, filter ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_name]

    # ── Row height header ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22

    # ── Example rows ──────────────────────────────────────────────────────────
    col_index = {name: idx for idx, name in enumerate(COLUMNS)}   # 0-based

    for row_idx, row_data in enumerate(EXAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            col_name = COLUMNS[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = EXAMPLE_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER

            # Apply number formats
            if col_name == "date":
                cell.number_format = FMT_DATE
            elif col_name == "volume":
                cell.number_format = FMT_VOLUME
            elif col_name in PRICE_COLS:
                cell.number_format = FMT_PRICE
            # template_version, ticker, timeframe → plain text (no special format)

    # Mark first example row with a comment on column A
    comment_text = (
        "EXAMPLE ROW — Replace with your own data.\n"
        "Keep all columns. Export sheet as CSV (UTF-8) before uploading."
    )
    ws.cell(row=2, column=1).comment = Comment(comment_text, "Stock Pattern Engine")

    # ── Data Validations ──────────────────────────────────────────────────────
    last_data_row = 10000   # apply validations to a large range

    # 1. timeframe dropdown
    timeframe_col = get_column_letter(COLUMNS.index("timeframe") + 1)
    dv_timeframe = DataValidation(
        type="list",
        formula1=f'"{",".join(TIMEFRAME_OPTIONS)}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Timeframe",
        error=f'Please select one of: {", ".join(TIMEFRAME_OPTIONS)}',
        showDropDown=False,   # False = show the dropdown arrow in Excel
    )
    dv_timeframe.sqref = f"{timeframe_col}2:{timeframe_col}{last_data_row}"
    ws.add_data_validation(dv_timeframe)

    # 2. volume >= 0
    volume_col = get_column_letter(COLUMNS.index("volume") + 1)
    dv_volume = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Volume",
        error="Volume must be a non-negative integer.",
    )
    dv_volume.sqref = f"{volume_col}2:{volume_col}{last_data_row}"
    ws.add_data_validation(dv_volume)

    # 3. RSI 0–100
    rsi_col = get_column_letter(COLUMNS.index("RSI") + 1)
    dv_rsi = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid RSI",
        error="RSI must be between 0 and 100.",
    )
    dv_rsi.sqref = f"{rsi_col}2:{rsi_col}{last_data_row}"
    ws.add_data_validation(dv_rsi)

    # 4. open/high/low/close must be positive decimals (Excel-level: > 0)
    for price_col_name in ("open", "high", "low", "close"):
        pc = get_column_letter(COLUMNS.index(price_col_name) + 1)
        dv_price = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=f"Invalid {price_col_name.title()}",
            error=f"{price_col_name.title()} price must be greater than 0.",
        )
        dv_price.sqref = f"{pc}2:{pc}{last_data_row}"
        ws.add_data_validation(dv_price)

    # Note: high>=open/close and low<=open/close require cross-column formulas.
    # These are enforced by the backend validator (core/validator.py OHLC sanity check).


# ─── Sheet 2: INSTRUCTIONS ────────────────────────────────────────────────────

COLUMN_DOCS = [
    ("template_version", "1.0.0",       "Do not change. Used by backend to detect template version."),
    ("ticker",           "AAPL",         "Stock symbol, uppercase letters only (e.g. AAPL, TSLA, NVDA)."),
    ("timeframe",        "1D",           "Bar timeframe. Must be one of: 1D, 4H, 1H, 15M."),
    ("date",             "2024-01-15",   "Bar date in YYYY-MM-DD format. Must be chronologically sorted."),
    ("open",             "185.00",       "Opening price of the bar. Positive decimal. No commas."),
    ("high",             "186.92",       "Highest price of the bar. Must be >= open and close."),
    ("low",              "183.38",       "Lowest price of the bar. Must be <= open and close."),
    ("close",            "185.92",       "Closing price of the bar. Positive decimal."),
    ("volume",           "52341800",     "Total shares/contracts traded. Non-negative integer."),
    ("RSI",              "54.30",        "14-period Relative Strength Index. Range: 0–100."),
    ("MACD",             "0.42",         "MACD line (EMA12 − EMA26). Can be negative."),
    ("MACD_signal",      "0.28",         "9-period EMA of MACD line."),
    ("MACD_histogram",   "0.14",         "MACD − Signal line."),
    ("EMA_20",           "183.10",       "20-period Exponential Moving Average."),
    ("EMA_50",           "179.50",       "50-period Exponential Moving Average."),
    ("SMA_200",          "168.20",       "200-period Simple Moving Average."),
    ("Bollinger_upper",  "190.10",       "Upper Bollinger Band (20-period SMA + 2σ)."),
    ("Bollinger_lower",  "175.80",       "Lower Bollinger Band (20-period SMA − 2σ)."),
    ("ATR",              "2.85",         "14-period Average True Range."),
]

WARNINGS = [
    "DO NOT rename any column headers. The backend rejects files with missing or renamed columns.",
    "DO NOT delete required columns: ticker, date, open, high, low, close, volume.",
    "DO NOT leave the header row (row 1) blank or shift it down.",
    "Indicator columns (RSI, MACD, etc.) are optional but improve signal quality.",
    "All dates must be in YYYY-MM-DD format (e.g. 2024-01-15, not 01/15/2024).",
    "High must be ≥ open AND close. Low must be ≤ open AND close.",
    "Export as CSV UTF-8 (File → Save As → CSV UTF-8) before uploading.",
]


def _build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 62

    row = 1

    # Title
    title = ws.cell(row=row, column=1, value="Stock Pattern Engine — Template Guide")
    title.font      = INSTRUCTION_TITLE
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 26
    row += 1

    ws.cell(row=row, column=1,
            value=f"Template Version: {TEMPLATE_VERSION}  |  Min rows required: 20  |  Max rows: 5,000"
    ).font = Font(name="Calibri", italic=True, size=10, color="595959")
    ws.merge_cells(f"A{row}:C{row}")
    row += 2

    # ── Column reference ──────────────────────────────────────────────────────
    hdr_row = row
    for col_idx, header in enumerate(["Column Name", "Example Value", "Description"], start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=header)
        cell.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = CENTER
    ws.row_dimensions[hdr_row].height = 18
    row += 1

    req_cols = {"ticker", "date", "open", "high", "low", "close", "volume"}
    for col_name, example, description in COLUMN_DOCS:
        is_required = col_name in req_cols
        fill = PatternFill("solid", fgColor="EBF3FB") if is_required else PatternFill("solid", fgColor="FFFFFF")

        c1 = ws.cell(row=row, column=1, value=col_name)
        c2 = ws.cell(row=row, column=2, value=example)
        c3 = ws.cell(row=row, column=3, value=("* REQUIRED — " if is_required else "  optional  — ") + description)

        for cell in (c1, c2, c3):
            cell.font      = Font(name="Calibri", size=10, bold=is_required)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # ── Warnings ──────────────────────────────────────────────────────────────
    warn_title = ws.cell(row=row, column=1, value="⚠  Important Rules")
    warn_title.font = Font(name="Calibri", bold=True, size=11, color="C00000")
    ws.merge_cells(f"A{row}:C{row}")
    row += 1

    for warning in WARNINGS:
        c = ws.cell(row=row, column=1, value=f"  •  {warning}")
        c.font      = Font(name="Calibri", size=10)
        c.alignment = WRAP
        c.border    = THIN_BORDER
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ── Export reminder ───────────────────────────────────────────────────────
    export = ws.cell(
        row=row, column=1,
        value="  How to export:  File → Save As → CSV UTF-8 (Comma delimited) (.csv)"
    )
    export.font      = Font(name="Calibri", bold=True, size=11, color="375623")
    export.fill      = PatternFill("solid", fgColor="E2EFDA")
    export.alignment = WRAP
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 22

    # ── Backend validation note ────────────────────────────────────────────────
    row += 2
    note = ws.cell(
        row=row, column=1,
        value=(
            "ℹ  Backend validation enforced (beyond Excel rules): "
            "high ≥ open & close; low ≤ open & close; "
            "date format YYYY-MM-DD; no duplicate dates per ticker; "
            "ticker uppercase; volume ≥ 0."
        )
    )
    note.font      = Font(name="Calibri", italic=True, size=9, color="595959")
    note.alignment = WRAP
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 30
