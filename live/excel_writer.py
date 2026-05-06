"""
Excel writer — keeps ~/Desktop/StockWatchlist.xlsx in sync with live data.
Python writes every column EXCEPT 'Purchase Price' (column B) which the
user fills in manually.
"""
import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXCEL_PATH = Path.home() / "Desktop" / "StockWatchlist.xlsx"

HEADERS = [
    "Ticker", "Purchase Price *", "Live Price", "Change %", "P&L %",
    "Pattern", "Confidence", "Signal", "Entry", "Stop Loss",
    "TP-1", "TP-2", "TP-3", "R/R Ratio", "Risk", "Last Updated",
]

# Column indices (1-based)
COL_TICKER          = 1
COL_PURCHASE_PRICE  = 2   # ← only manual column
COL_LIVE_PRICE      = 3
COL_CHANGE_PCT      = 4
COL_PNL             = 5
COL_PATTERN         = 6
COL_CONFIDENCE      = 7
COL_SIGNAL          = 8
COL_ENTRY           = 9
COL_STOP            = 10
COL_TP1             = 11
COL_TP2             = 12
COL_TP3             = 13
COL_RR              = 14
COL_RISK            = 15
COL_UPDATED         = 16

COL_WIDTHS = [
    8, 16, 12, 10, 10, 24, 12, 10, 10, 10, 10, 10, 10, 10, 10, 18
]

# Colour palette
C_NAVY   = "1F3864"
C_WHITE  = "FFFFFF"
C_GREEN  = "15803D"
C_RED    = "B91C1C"
C_AMBER  = "B45309"
C_LGRAY  = "F1F5F9"
C_BLUE   = "EBF3FB"
C_MED    = "E2E8F0"
C_MUTED  = "94A3B8"


def _thin():
    return Border(bottom=Side(style="thin", color=C_MED))


def _hdr_row(ws, row=1):
    fill = PatternFill("solid", fgColor=C_NAVY)
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Public API ─────────────────────────────────────────────────────────────────

def read_purchase_prices() -> dict:
    """Read user-entered purchase prices from column B. Safe to call even if
    file doesn't exist yet."""
    if not EXCEL_PATH.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        prices = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ticker = str(row[0]).upper()
                try:
                    prices[ticker] = float(row[1]) if row[1] is not None else None
                except (TypeError, ValueError):
                    prices[ticker] = None
        return prices
    except Exception as e:
        logger.warning(f"read_purchase_prices: {e}")
        return {}


def write_live_data(live_results: list):
    """
    Rewrite the Excel file with fresh live data.
    Preserves user-entered purchase prices from column B.
    """
    purchase_prices = read_purchase_prices()

    try:
        if EXCEL_PATH.exists():
            wb = openpyxl.load_workbook(EXCEL_PATH)
        else:
            wb = openpyxl.Workbook()
    except Exception:
        wb = openpyxl.Workbook()

    sheet_name = "Live Watchlist"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row + 5)
    else:
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet(sheet_name)

    ws.freeze_panes = "A2"
    _hdr_row(ws)
    _set_col_widths(ws)

    for i, res in enumerate(live_results, start=2):
        ticker     = res.get("ticker", "")
        pp         = purchase_prices.get(ticker) or res.get("purchase_price")
        live_price = res.get("live_price")
        change_pct = res.get("change_pct")

        pnl = None
        if pp and live_price:
            pnl = round((live_price - pp) / pp * 100, 2)

        row_vals = [
            ticker, pp, live_price, change_pct, pnl,
            res.get("pattern", "—"),
            res.get("confidence"),
            res.get("signal", "—"),
            res.get("entry"),
            res.get("stop_loss"),
            res.get("tp1"),
            res.get("tp2"),
            res.get("tp3"),
            res.get("rr_ratio"),
            res.get("risk", "—"),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ]

        bg = C_LGRAY if i % 2 == 0 else C_WHITE
        row_fill = PatternFill("solid", fgColor=bg)

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=9)
            cell.fill = row_fill
            cell.border = _thin()

            # Purchase price column — highlight as editable
            if col == COL_PURCHASE_PRICE:
                cell.fill = PatternFill("solid", fgColor=C_BLUE)
                cell.font = Font(name="Calibri", size=9, bold=True)
                cell.number_format = "$#,##0.00"

            # Price columns
            if col in (COL_LIVE_PRICE, COL_ENTRY, COL_STOP, COL_TP1, COL_TP2, COL_TP3):
                cell.number_format = "$#,##0.00"

            # Percentage columns
            if col == COL_CHANGE_PCT and change_pct is not None:
                cell.number_format = "0.00%"
                clr = C_GREEN if change_pct >= 0 else C_RED
                cell.font = Font(name="Calibri", size=9, bold=True, color=clr)

            if col == COL_PNL and pnl is not None:
                cell.number_format = "0.00%"
                clr = C_GREEN if pnl >= 0 else C_RED
                cell.font = Font(name="Calibri", size=9, bold=True, color=clr)

            # Signal colour
            if col == COL_SIGNAL:
                clr = {"BUY": C_GREEN, "SELL": C_RED, "WATCH": C_AMBER}.get(
                    str(val).upper(), "475569"
                )
                cell.font = Font(name="Calibri", size=9, bold=True, color=clr)

            # Risk colour
            if col == COL_RISK:
                clr = {"LOW": C_GREEN, "HIGH": C_RED, "MEDIUM": C_AMBER}.get(
                    str(val).upper(), "475569"
                )
                cell.font = Font(name="Calibri", size=9, bold=True, color=clr)

        ws.row_dimensions[i].height = 18

    # Footer note
    note_row = len(live_results) + 3
    cell = ws.cell(
        row=note_row, column=1,
        value="* Purchase Price: Enter manually in column B. "
              "All other columns update automatically every 5 minutes during market hours.",
    )
    cell.font = Font(name="Calibri", italic=True, size=8, color=C_MUTED)
    ws.merge_cells(f"A{note_row}:{get_column_letter(len(HEADERS))}{note_row}")

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(EXCEL_PATH)
        logger.info(f"Excel updated → {EXCEL_PATH}")
    except PermissionError:
        logger.warning(
            f"Excel file is open in Excel — close it to allow auto-updates: {EXCEL_PATH}"
        )
    except Exception as e:
        logger.error(f"Excel write error: {e}")
